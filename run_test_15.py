import openpyxl
from openpyxl.styles import PatternFill
from amazon_scraper import AmazonScraper
import time
import os

# Define colors for conditional formatting
GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

def run_test_15():
    input_file = "/Users/leibykoplowitz/Downloads/2025 master maintenance.xlsx"
    output_file = "/Users/leibykoplowitz/Downloads/2025 master maintenance_test_15.xlsx"
    
    print(f"Starting test run on 15 products...")
    print(f"Input: {input_file}")
    print(f"Output: {output_file}")
    
    # Load workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # Initialize scraper (headless)
    scraper = AmazonScraper(headless=True)
    
    try:
        # Process rows 3 to 17 (15 products)
        for row_num in range(3, 18):
            asin = ws.cell(row=row_num, column=2).value
            if not asin:
                continue
            
            # Clean ASIN
            asin = str(asin).strip()
            expected_price = ws.cell(row=row_num, column=3).value
            
            print(f"\n{'='*60}")
            print(f"Processing Row {row_num}: {asin}")
            print(f"{'='*60}")
            
            # Scrape data
            data = scraper.scrape_product(asin, expected_price)
            
            if data:
                # Update Excel with scraped data
                
                # Column F: Link
                ws.cell(row_num, 6).value = data['link']
                
                # Column G: BuyBox Seller
                buybox_seller = data['buybox_seller']
                if buybox_seller and 'amazon.com' in buybox_seller.lower():
                    ws.cell(row_num, 7).value = "YES"
                    ws.cell(row_num, 7).fill = GREEN_FILL
                else:
                    ws.cell(row_num, 7).value = buybox_seller or "Unknown"
                    ws.cell(row_num, 7).fill = RED_FILL
                
                # Column H: Price
                buybox_price = data['buybox_price']
                if buybox_price is not None:
                    ws.cell(row_num, 8).value = buybox_price
                    if expected_price and abs(buybox_price - expected_price) < 0.01:
                        ws.cell(row_num, 8).fill = GREEN_FILL
                    else:
                        ws.cell(row_num, 8).fill = RED_FILL
                
                # Column I: Ranking
                ws.cell(row_num, 9).value = data['ranking']
                
                # Column J: Review
                ws.cell(row_num, 10).value = data['review']
                
                # Column K: Photos
                photo_count = data['photos']
                if photo_count >= 8:
                    ws.cell(row_num, 11).value = "GOOD"
                    ws.cell(row_num, 11).fill = GREEN_FILL
                else:
                    ws.cell(row_num, 11).value = photo_count
                    ws.cell(row_num, 11).fill = RED_FILL
                
                # Column L: Videos
                ws.cell(row_num, 12).value = data['videos']
                
                # Column M: Bullet Points
                ws.cell(row_num, 13).value = data['bullet_points']
                
                print(f"✓ Successfully processed {asin}")
                print(f"  Photos: {data['photos']}")
                print(f"  Ranking: {data['ranking']}")
            else:
                print(f"✗ Failed to scrape {asin}")
            
            # Save progress
            wb.save(output_file)
            
            # Small delay to be polite
            time.sleep(2)
            
    except Exception as e:
        print(f"\n❌ Error during test run: {e}")
    finally:
        scraper.close()
        print(f"\n{'='*60}")
        print("TEST COMPLETE!")
        print(f"Output saved to: {output_file}")
        print(f"{'='*60}")

if __name__ == "__main__":
    run_test_15()
