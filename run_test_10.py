#!/usr/bin/env python3
"""
Run Amazon scraper on first 10 products as a test
"""

import openpyxl
from openpyxl.styles import PatternFill
from amazon_scraper import AmazonScraper
import time

# Color fills for Excel
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

input_file = "/Users/leibykoplowitz/Downloads/2025 master maintenance.xlsx"
output_file = "/Users/leibykoplowitz/Downloads/2025 master maintenance_test_10.xlsx"

print("="*60)
print("Amazon Scraper - Test Run (10 Products)")
print("="*60)
print(f"Input:  {input_file}")
print(f"Output: {output_file}")
print("="*60)

# Load workbook
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# Initialize scraper (headless mode for faster execution)
scraper = AmazonScraper(headless=True)

try:
    # Process rows 3-12 (first 10 products after the example)
    for row_num in range(3, 13):
        asin = ws.cell(row_num, 2).value  # Column B
        expected_price = ws.cell(row_num, 3).value  # Column C
        
        if not asin:
            print(f"\nRow {row_num}: Skipping (no ASIN)")
            continue
        
        # Clean ASIN (remove spaces)
        asin = str(asin).strip()
        
        print(f"\n{'='*60}")
        print(f"Processing Row {row_num}/{12}: {asin}")
        print(f"{'='*60}")
        
        # Scrape product data
        data = scraper.scrape_product(asin, expected_price)
        
        if data:
            # Column F: Link
            ws.cell(row_num, 6).value = data['link']
            
            # Column G: BuyBox Seller
            buybox_seller = data['buybox_seller']
            if buybox_seller and 'amazon.com' in buybox_seller.lower():
                ws.cell(row_num, 7).value = "YES"
                ws.cell(row_num, 7).fill = GREEN_FILL
                print(f"  ✓ Buybox: Amazon.com (GREEN)")
            else:
                ws.cell(row_num, 7).value = buybox_seller or "Unknown"
                ws.cell(row_num, 7).fill = RED_FILL
                print(f"  ✗ Buybox: {buybox_seller} (RED)")
            
            # Column H: Price
            buybox_price = data['buybox_price']
            if buybox_price is not None:
                ws.cell(row_num, 8).value = buybox_price
                if expected_price and abs(buybox_price - expected_price) < 0.01:
                    ws.cell(row_num, 8).fill = GREEN_FILL
                    print(f"  ✓ Price: ${buybox_price} matches ${expected_price} (GREEN)")
                else:
                    ws.cell(row_num, 8).fill = RED_FILL
                    print(f"  ✗ Price: ${buybox_price} vs expected ${expected_price} (RED)")
            
            # Column I: Ranking
            ws.cell(row_num, 9).value = data['ranking']
            print(f"  • Ranking: {data['ranking']}")
            
            # Column J: Review
            ws.cell(row_num, 10).value = data['review']
            print(f"  • Review: {data['review']}")
            
            # Column K: Photos
            photo_count = data['photos']
            if photo_count >= 8:
                ws.cell(row_num, 11).value = "GOOD"
                ws.cell(row_num, 11).fill = GREEN_FILL
                print(f"  ✓ Photos: {photo_count} (GOOD - GREEN)")
            else:
                ws.cell(row_num, 11).value = photo_count
                ws.cell(row_num, 11).fill = RED_FILL
                print(f"  ✗ Photos: {photo_count} (needs 8+ - RED)")
            
            # Column L: Videos
            ws.cell(row_num, 12).value = data['videos']
            print(f"  • Videos: {data['videos']}")
            
            # Column M: Bullet Points
            ws.cell(row_num, 13).value = data['bullet_points']
            print(f"  • Bullet Points: {data['bullet_points']}")
            
            print(f"\n  ✓ Successfully processed {asin}")
        else:
            print(f"  ✗ Failed to scrape {asin}")
        
        # Save progress after each row
        wb.save(output_file)
        
        # Small delay between requests
        time.sleep(2)

finally:
    scraper.close()
    wb.save(output_file)
    print(f"\n{'='*60}")
    print(f"TEST COMPLETE!")
    print(f"Output saved to: {output_file}")
    print(f"{'='*60}")
