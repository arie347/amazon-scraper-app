from fastapi import FastAPI, Request, UploadFile, File, BackgroundTasks
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import shutil
import os
import uuid
import database
from amazon_scraper import AmazonScraper
import openpyxl
from openpyxl.styles import PatternFill

app = FastAPI()

# Setup directories
UPLOAD_DIR = "uploads"
RESULTS_DIR = "results"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(RESULTS_DIR, exist_ok=True)

# Setup templates
templates = Jinja2Templates(directory="templates")

# Initialize DB
database.init_db()

# Define colors (copied from script)
GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

def run_scraper_task(file_id: int, input_path: str):
    """Background task to run the scraper"""
    try:
        database.update_status(file_id, "Running")
        
        # Define output path
        filename = os.path.basename(input_path)
        result_filename = f"updated_{filename}"
        output_path = os.path.join(RESULTS_DIR, result_filename)
        
        # Load workbook
        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        
        # Initialize scraper
        scraper = AmazonScraper(headless=True)
        
        try:
            # Count total rows to process
            total_rows = 0
            for row_num in range(3, ws.max_row + 1):
                if ws.cell(row_num, 2).value:
                    total_rows += 1
            
            database.update_progress(file_id, 0, total_rows)
            
            # Process rows
            current_row = 0
            for row_num in range(3, ws.max_row + 1):
                asin = ws.cell(row_num, 2).value
                if not asin:
                    continue
                
                asin = str(asin).strip()
                expected_price = ws.cell(row_num, 3).value
                
                # Scrape
                data = scraper.scrape_product(asin, expected_price)
                
                # Update progress
                current_row += 1
                database.update_progress(file_id, current_row, total_rows)
                
                if data:
                    # Update Excel (Logic copied from run_test_15.py)
                    ws.cell(row_num, 6).value = data['link']
                    
                    buybox_seller = data['buybox_seller']
                    if buybox_seller and 'amazon.com' in buybox_seller.lower():
                        ws.cell(row_num, 7).value = "YES"
                        ws.cell(row_num, 7).fill = GREEN_FILL
                    else:
                        ws.cell(row_num, 7).value = buybox_seller or "Unknown"
                        ws.cell(row_num, 7).fill = RED_FILL
                    
                    buybox_price = data['buybox_price']
                    if buybox_price is not None:
                        ws.cell(row_num, 8).value = buybox_price
                        if expected_price and abs(buybox_price - expected_price) < 0.01:
                            ws.cell(row_num, 8).fill = GREEN_FILL
                        else:
                            ws.cell(row_num, 8).fill = RED_FILL
                    
                    ws.cell(row_num, 9).value = data['ranking']
                    ws.cell(row_num, 10).value = data['review']
                    
                    photo_count = data['photos']
                    if photo_count >= 8:
                        ws.cell(row_num, 11).value = "GOOD"
                        ws.cell(row_num, 11).fill = GREEN_FILL
                    else:
                        ws.cell(row_num, 11).value = photo_count
                        ws.cell(row_num, 11).fill = RED_FILL
                    
                    ws.cell(row_num, 12).value = data['videos']
                    ws.cell(row_num, 13).value = data['bullet_points']
            
            # Save result
            wb.save(output_path)
            database.update_status(file_id, "Completed", result_filename)
            
        except Exception as e:
            print(f"Scraping error: {e}")
            database.update_status(file_id, "Failed")
        finally:
            scraper.close()
            
    except Exception as e:
        print(f"Task error: {e}")
        database.update_status(file_id, "Failed")

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    files = database.get_all_files()
    return templates.TemplateResponse("index.html", {"request": request, "files": files})

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    # Generate safe filename
    safe_filename = f"{uuid.uuid4()}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIR, safe_filename)
    
    # Save file
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Add to DB
    database.add_file(safe_filename, file.filename)
    
    return RedirectResponse(url="/", status_code=303)

@app.post("/scrape/{file_id}")
async def start_scrape(file_id: int, background_tasks: BackgroundTasks):
    file_info = database.get_file(file_id)
    if file_info:
        input_path = os.path.join(UPLOAD_DIR, file_info['filename'])
        background_tasks.add_task(run_scraper_task, file_id, input_path)
    return RedirectResponse(url="/", status_code=303)

@app.get("/download/{file_id}")
async def download_result(file_id: int):
    file_info = database.get_file(file_id)
    if file_info and file_info['result_filename']:
        path = os.path.join(RESULTS_DIR, file_info['result_filename'])
        return FileResponse(path, filename=f"UPDATED_{file_info['original_filename']}")
    return RedirectResponse(url="/")

@app.post("/delete/{file_id}")
async def delete_file(file_id: int):
    file_info = database.get_file(file_id)
    if file_info:
        # Try to delete actual files
        try:
            os.remove(os.path.join(UPLOAD_DIR, file_info['filename']))
            if file_info['result_filename']:
                os.remove(os.path.join(RESULTS_DIR, file_info['result_filename']))
        except:
            pass
        
        database.delete_file(file_id)
    return RedirectResponse(url="/", status_code=303)
