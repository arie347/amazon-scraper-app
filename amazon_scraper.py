#!/usr/bin/env python3
"""
Amazon Product Scraper
Scrapes Amazon product listings and fills out Excel sheet with validation
"""

import openpyxl
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import time
import re

# Color fills for Excel
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

class AmazonScraper:
    def __init__(self, headless=False):
        """Initialize the scraper with Chrome driver"""
        options = webdriver.ChromeOptions()
        if headless:
            options.add_argument('--headless')  # Run in background
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        self.driver = webdriver.Chrome(options=options)
        self.wait = WebDriverWait(self.driver, 10)
    
    def get_product_url(self, asin):
        """Generate Amazon product URL from ASIN"""
        return f"https://www.amazon.com/dp/{asin}"
    
    def scrape_product(self, asin, expected_price):
        """Scrape all required information for a product"""
        url = self.get_product_url(asin)
        print(f"\nScraping ASIN: {asin}")
        print(f"URL: {url}")
        
        try:
            self.driver.get(url)
            time.sleep(2)  # Wait for page to load
            
            data = {
                'link': url,
                'buybox_seller': None,
                'buybox_price': None,
                'ranking': None,
                'review': None,
                'photos': None,
                'videos': None,
                'bullet_points': None
            }
            
            # Extract buybox seller
            data['buybox_seller'] = self._get_buybox_seller()
            
            # Extract price
            data['buybox_price'] = self._get_price()
            
            # Extract ranking
            data['ranking'] = self._get_ranking()
            
            # Extract review rating
            data['review'] = self._get_review_rating()
            
            # Count photos
            data['photos'] = self._count_photos()
            
            # Check for videos
            data['videos'] = self._check_videos()
            
            # Count bullet points
            data['bullet_points'] = self._count_bullet_points()
            
            return data
            
        except Exception as e:
            print(f"Error scraping {asin}: {str(e)}")
            return None
    
    def _get_buybox_seller(self):
        """Extract the buybox seller name"""
        try:
            # Method 1: Check for "Ships from and sold by Amazon.com"
            try:
                merchant_info = self.driver.find_element(By.ID, "merchant-info")
                text = merchant_info.text.lower()
                if 'amazon.com' in text or 'amazon' in text:
                    return "Amazon.com"
                # Extract seller name from link
                try:
                    seller_link = merchant_info.find_element(By.TAG_NAME, "a")
                    return seller_link.text.strip()
                except:
                    pass
            except:
                pass
            
            # Method 2: Check tabular buybox
            try:
                seller_element = self.driver.find_element(By.ID, "sellerProfileTriggerId")
                seller_name = seller_element.text.strip()
                if seller_name:
                    return seller_name
            except:
                pass
            
            # Method 3: Look for "Sold by" text anywhere
            try:
                sold_by_elements = self.driver.find_elements(By.XPATH, "//*[contains(text(), 'Sold by')]")
                for elem in sold_by_elements:
                    parent_text = elem.find_element(By.XPATH, "./..").text
                    if 'amazon.com' in parent_text.lower() or 'amazon' in parent_text.lower():
                        return "Amazon.com"
                    # Try to extract seller name
                    try:
                        seller_link = elem.find_element(By.XPATH, ".//following-sibling::a | .//a")
                        return seller_link.text.strip()
                    except:
                        pass
            except:
                pass
            
            # Method 4: Check if "Add to Cart" button exists (usually means Amazon is seller)
            try:
                add_to_cart = self.driver.find_element(By.ID, "add-to-cart-button")
                if add_to_cart:
                    # If we can add to cart but haven't found seller, likely Amazon
                    return "Amazon.com"
            except:
                pass
            
            return "Unknown"
        except Exception as e:
            print(f"Error getting buybox seller: {e}")
            return "Unknown"
    
    def _get_price(self):
        """Extract the current buybox price"""
        try:
            # Method 1: Try to get the offscreen price (most reliable, contains full price)
            try:
                price_element = self.driver.find_element(By.CSS_SELECTOR, ".a-price .a-offscreen")
                price_text = price_element.get_attribute("textContent")
                # Extract numeric value (e.g., "$89.99" -> 89.99)
                price_match = re.search(r'[\d,]+\.?\d*', price_text)
                if price_match:
                    price = float(price_match.group().replace(',', ''))
                    print(f"Found price (offscreen): ${price}")
                    return price
            except:
                pass
            
            # Method 2: Combine whole and fraction parts
            try:
                whole_element = self.driver.find_element(By.CLASS_NAME, "a-price-whole")
                whole_text = whole_element.text.strip().replace(',', '').replace('.', '')  # Remove comma and trailing dot
                
                # Try to get the fraction part
                fraction_text = "00"
                try:
                    fraction_element = self.driver.find_element(By.CLASS_NAME, "a-price-fraction")
                    fraction_text = fraction_element.text.strip()
                except:
                    pass
                
                # Combine whole and fraction
                price = float(f"{whole_text}.{fraction_text}")
                print(f"Found price (whole+fraction): ${price}")
                return price
            except:
                pass
            
            # Method 3: Old price selectors (fallback)
            price_selectors = [
                (By.ID, "priceblock_ourprice"),
                (By.ID, "priceblock_dealprice"),
            ]
            
            for selector_type, selector in price_selectors:
                try:
                    price_element = self.driver.find_element(selector_type, selector)
                    price_text = price_element.text
                    # Extract numeric value
                    price_match = re.search(r'[\d,]+\.?\d*', price_text)
                    if price_match:
                        price = float(price_match.group().replace(',', ''))
                        print(f"Found price (fallback): ${price}")
                        return price
                except:
                    continue
            
            print("No price found")
            return None
        except Exception as e:
            print(f"Error getting price: {e}")
            return None
    
    def _get_ranking(self):
        """Extract the Amazon Best Sellers Rank (lowest number)"""
        try:
            # Look for the ranking text
            # It's usually in a list item with id containing 'SalesRank' or just text in the product details
            
            rankings = []
            
            # Method 1: Look for specific SalesRank list items
            try:
                # This selector often finds the main ranking block
                elements = self.driver.find_elements(By.XPATH, "//*[contains(text(), 'Best Sellers Rank')]/parent::*")
                for elem in elements:
                    text = elem.text
                    # Extract all rankings from this block
                    # Pattern: #123 in Category
                    matches = re.finditer(r'#([\d,]+)\s+in\s+([^\(\n]+)', text)
                    for match in matches:
                        rank_num = int(match.group(1).replace(',', ''))
                        full_text = f"#{match.group(1)} in {match.group(2).strip()}"
                        rankings.append((rank_num, full_text))
            except:
                pass
                
            # Method 2: Look for any text containing "#" and "in" (fallback)
            if not rankings:
                try:
                    body_text = self.driver.find_element(By.TAG_NAME, "body").text
                    # Limit search to likely areas or just regex the whole body if needed (can be slow/noisy)
                    # Better to look for specific containers if possible, but let's try a targeted regex on the page source
                    # actually, let's stick to specific elements to avoid noise
                    pass
                except:
                    pass

            if rankings:
                # Sort by rank number (ascending) to get the best rank
                rankings.sort(key=lambda x: x[0])
                best_rank = rankings[0][1]
                print(f"Found rankings: {rankings}")
                print(f"Selected best ranking: {best_rank}")
                return best_rank
            
            print("No ranking found")
            return None
        except Exception as e:
            print(f"Error getting ranking: {e}")
            return None
    
    def _get_review_rating(self):
        """Extract average review rating"""
        try:
            rating_selectors = [
                (By.CSS_SELECTOR, "span[data-hook='rating-out-of-text']"),
                (By.CSS_SELECTOR, "i[data-hook='average-star-rating'] span.a-icon-alt"),
                (By.XPATH, "//span[@id='acrPopover']/@title"),
            ]
            
            for selector_type, selector in rating_selectors:
                try:
                    rating_element = self.driver.find_element(selector_type, selector)
                    rating_text = rating_element.text if selector_type != By.XPATH else rating_element.get_attribute("title")
                    # Extract numeric rating (e.g., "4.5 out of 5 stars" -> 4.5)
                    rating_match = re.search(r'([\d.]+)\s*out of', rating_text)
                    if rating_match:
                        return float(rating_match.group(1))
                except:
                    continue
            
            return None
        except Exception as e:
            print(f"Error getting review rating: {e}")
            return None
    
    def _count_photos(self):
        """Count the number of product photos"""
        try:
            # Look for image thumbnails in the image block
            image_count = 0
            
            # Method 1: Count thumbnail images (most reliable)
            try:
                thumbnails = self.driver.find_elements(By.CSS_SELECTOR, "#altImages ul li.imageThumbnail")
                if thumbnails:
                    image_count = len(thumbnails)
                    
                    # Check for "+X" overlay on the last thumbnail
                    # The overlay is often in a span following the input, or part of the button label
                    try:
                        # Look for text like "5+" or "+5" in the altImages section
                        overlay_elements = self.driver.find_elements(By.CSS_SELECTOR, "#altImages .a-button-text span")
                        for elem in overlay_elements:
                            text = elem.text.strip()
                            # Check for pattern like "5+" or "+5"
                            match = re.search(r'(\d+)\+', text) or re.search(r'\+(\d+)', text)
                            if match:
                                hidden_count = int(match.group(1))
                                print(f"Found hidden images overlay: {text} (adding {hidden_count})")
                                image_count += hidden_count
                                break
                    except:
                        pass
                        
                    print(f"Found {image_count} images via thumbnails")
                    return image_count
            except:
                pass
            
            # Method 2: Count all items in altImages (includes videos)
            try:
                all_items = self.driver.find_elements(By.CSS_SELECTOR, "#altImages ul li")
                video_items = self.driver.find_elements(By.CSS_SELECTOR, "#altImages ul li.videoThumbnail")
                image_count = len(all_items) - len(video_items)
                if image_count > 0:
                    print(f"Found {image_count} images (total items: {len(all_items)}, videos: {len(video_items)})")
                    return image_count
            except:
                pass
            
            # Method 3: Alternative selector
            try:
                images = self.driver.find_elements(By.CSS_SELECTOR, "#imageBlock img")
                if images:
                    image_count = len(images)
                    print(f"Found {image_count} images via imageBlock")
                    return image_count
            except:
                pass
            
            print("No images found")
            return 0
        except Exception as e:
            print(f"Error counting photos: {e}")
            return 0
    
    def _check_videos(self):
        """Check if product has videos (excluding review videos)"""
        try:
            # Look for video elements in the image block
            video_selectors = [
                (By.CSS_SELECTOR, "#altImages ul li.videoThumbnail"),
                (By.CSS_SELECTOR, "li[data-csa-c-type='video']"),
                (By.XPATH, "//li[contains(@class, 'video')]"),
            ]
            
            for selector_type, selector in video_selectors:
                try:
                    videos = self.driver.find_elements(selector_type, selector)
                    if videos and len(videos) > 0:
                        return "YES"
                except:
                    continue
            
            return "NO"
        except Exception as e:
            print(f"Error checking videos: {e}")
            return "NO"
    
    def _count_bullet_points(self):
        """Count the number of bullet points in product description"""
        try:
            bullet_selectors = [
                (By.CSS_SELECTOR, "#feature-bullets ul li"),
                (By.XPATH, "//div[@id='feature-bullets']//li"),
            ]
            
            for selector_type, selector in bullet_selectors:
                try:
                    bullets = self.driver.find_elements(selector_type, selector)
                    # Filter out empty bullets
                    valid_bullets = [b for b in bullets if b.text.strip()]
                    if valid_bullets:
                        count = len(valid_bullets)
                        return "YES" if count >= 5 else "NO"
                except:
                    continue
            
            return "NO"
        except Exception as e:
            print(f"Error counting bullet points: {e}")
            return "NO"
    
    def close(self):
        """Close the browser"""
        self.driver.quit()


def process_excel(file_path, output_path=None):
    """Process the Excel file and fill in the scraped data"""
    if output_path is None:
        output_path = file_path.replace('.xlsx', '_updated.xlsx')
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Initialize scraper
    scraper = AmazonScraper(headless=True)
    
    try:
        # Process each row (starting from row 3, since row 2 is the example)
        for row_num in range(3, ws.max_row + 1):
            asin = ws.cell(row_num, 2).value  # Column B
            expected_price = ws.cell(row_num, 3).value  # Column C
            
            if not asin:
                continue
            
            # Clean ASIN (remove spaces)
            asin = str(asin).strip()
            
            print(f"\n{'='*60}")
            print(f"Processing Row {row_num}: {asin}")
            print(f"{'='*60}")
            
            # Scrape product data
            data = scraper.scrape_product(asin, expected_price)
            
            if data:
                # Column F: Link
                ws.cell(row_num, 6).value = data['link']
                
                # Column G: BuyBox Seller
                buybox_seller = data['buybox_seller']
                if buybox_seller and 'amazon.com' in buybox_seller.lower():
                    ws.cell(row_num, 7).value = "YES"
                    ws.cell(row_num, 7).fill = GREEN_FILL
                else:
                    ws.cell(row_num, 7).value = buybox_seller or "Unknown"
                    ws.cell(row_num, 7).fill = RED_FILL
                
                # Column H: Price
                buybox_price = data['buybox_price']
                if buybox_price is not None:
                    ws.cell(row_num, 8).value = buybox_price
                    if expected_price and abs(buybox_price - expected_price) < 0.01:
                        ws.cell(row_num, 8).fill = GREEN_FILL
                    else:
                        ws.cell(row_num, 8).fill = RED_FILL
                
                # Column I: Ranking
                ws.cell(row_num, 9).value = data['ranking']
                
                # Column J: Review
                ws.cell(row_num, 10).value = data['review']
                
                # Column K: Photos
                photo_count = data['photos']
                if photo_count >= 8:
                    ws.cell(row_num, 11).value = "GOOD"
                    ws.cell(row_num, 11).fill = GREEN_FILL
                else:
                    ws.cell(row_num, 11).value = photo_count
                    ws.cell(row_num, 11).fill = RED_FILL
                
                # Column L: Videos
                ws.cell(row_num, 12).value = data['videos']
                
                # Column M: Bullet Points
                ws.cell(row_num, 13).value = data['bullet_points']
                
                print(f"✓ Successfully processed {asin}")
            else:
                print(f"✗ Failed to scrape {asin}")
            
            # Save progress after each row
            wb.save(output_path)
            print(f"Progress saved to {output_path}")
            
            # Small delay between requests
            time.sleep(2)
    
    finally:
        scraper.close()
        wb.save(output_path)
        print(f"\n{'='*60}")
        print(f"Complete! Output saved to: {output_path}")
        print(f"{'='*60}")


if __name__ == "__main__":
    import sys
    
    input_file = "/Users/leibykoplowitz/Downloads/2025 master maintenance.xlsx"
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    
    output_file = input_file.replace('.xlsx', '_updated.xlsx')
    
    print(f"Starting Amazon scraper...")
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    
    process_excel(input_file, output_file)
